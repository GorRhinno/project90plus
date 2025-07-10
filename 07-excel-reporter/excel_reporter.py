from openpyxl import load_workbook
from collections import Counter

def excel_reporter(path: str, sheets: list[str]) -> None:
    wb = load_workbook(path)

    for sheet_name in sheets:
        ws = wb[sheet_name]
        total_salary = round(sum(float(row[0]) for row in ws.iter_rows(min_row=2, min_col=3, max_col=3, values_only=True)))
        average_salary = total_salary / len([row for row in ws.iter_rows(min_row=2, min_col=3, max_col=3, values_only=True)])
        department_counts = {}
        for row in ws.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):
            department = row[0]
            department_counts[department] = department_counts.get(department, 0) + 1
        #departments = [row[1] for row in rows]
        #department_counts = Counter(departments)
    summary = wb.create_sheet("Summary")
    summary.append(["Metric", "Value"])
    summary.append(["Total salary", total_salary])
    summary.append(["Average salary", average_salary])
    for department, count in department_counts.items():
        summary.append([f"{department} Employees", count])
    
    wb.save(path)
    
if __name__ == "__main__":
    excel_reporter("data.xlsx", ["Лист1"])