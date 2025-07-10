from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import pandas as pd

def main():
    work_book = load_workbook("yerevan_sales.xlsx")
    work_sheet = work_book.active
    for cell in work_sheet[1]:
        cell.font = Font(bold=True)
    yellow_fill = PatternFill("solid", fgColor="FFFF00")
    for row in work_sheet.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            cell.fill = yellow_fill
    red_font = Font(color="FF0000")
    for row in work_sheet.iter_rows(min_row=2):
        if row[0].value == "Итого":
            for cell in row:
                cell.font = red_font
    work_book.save("yerevan_customised.xlsx")
if __name__ == "__main__":
    main()